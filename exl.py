# Excelクラス
import openpyxl as px
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from pathlib import Path

class exl_wb:

    def __init__(self, foldername, filename):
        self.fld_pth = Path(foldername)
        self.fle_pth = self.fld_pth.joinpath(filename)
    
    # フォント・背景色変更
    def color_set(self, range, c_font, c_fg):
        for row in self.ws[range]:
            for cell in row:
                cell.font = c_font
                cell.fill = c_fg
    
    # 罫線描画
    def draw_line(self,range, c_bd):
        for row in self.ws[range]:
            for cell in row:
                cell.border = c_bd

    # メイン処理            
    def open_workbook(self, count=10):
        # 既存ファイル削除
        if self.fle_pth.exists():
            # self.wb = px.load_workbook(self.fle_pth)
            try:
                self.fle_pth.unlink()
            except:

                print('ファイル削除エラー')

        # 書式シート作成
        self.wb = px.Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = 'Ranking'
        self.ws.cell(row=2, column=2, value='対象URL')

        self.ws.cell(row=5, column=2, value='トップ')
        self.ws.cell(row=5, column=3, value='第１カテゴリ')
        self.ws.cell(row=5, column=4, value='第２カテゴリ')
        self.ws.cell(row=5, column=5, value='第３カテゴリ')
        self.ws.cell(row=5, column=6, value='第４カテゴリ')
        self.ws.cell(row=5, column=7, value='第５カテゴリ')

        self.ws.cell(row=8, column=2, value='順位')
        self.ws.cell(row=8, column=3, value='レビュー数')
        self.ws.cell(row=8, column=4, value='アベレージ')
        self.ws.cell(row=8, column=5, value='単価')
        self.ws.cell(row=8, column=6, value='商品名')
        self.ws.cell(row=8, column=7, value='画像URL')

        for i in range(count):
            target = self.ws.cell(row=9+i, column=2, value=f'{1+i}位')
            target.alignment = Alignment(horizontal='center', wrap_text=False)
        
        # フォント・背景色の指定
        ft = Font(color='FFFFFF')
        fg = PatternFill(fgColor='C05200', fill_type='solid')
        self.color_set("B5:G5", ft, fg)
        fg = PatternFill(fgColor='277E3E', fill_type='solid')
        self.color_set("B8:G8", ft, fg)
        fg = PatternFill(fgColor='2F929A', fill_type='solid')
        ft = Font(color='000000')
        self.color_set(f"B9:B{count+8}", ft, fg)

        # 列幅の指定
        self.ws.column_dimensions['C'].width = 15.0
        self.ws.column_dimensions['D'].width = 15.0
        self.ws.column_dimensions['E'].width = 15.0
        self.ws.column_dimensions['F'].width = 60.0
        self.ws.column_dimensions['G'].width = 60.0

        # 書式の指定
        for i in range(30):
            self.ws.cell(row=9+i, column=4).number_format = '#0.00'
            self.ws.cell(row=9+i, column=5).number_format = '#,##0'

        # 罫線の指定
        bd = Border(top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        )
        self.draw_line("B5:G6",bd)
        self.draw_line(f"B8:G{count+8}",bd)

        # 表示倍率の指定
        self.ws.sheet_view.zoomScale = 150
        
    # 書式シートのコピー
    def copy_ws(self):
        return self.wb.copy_worksheet(self.ws)

    # 終了処理
    def save_workbook(self):

        # 書式シート削除
        self.wb.remove(self.wb['Ranking'])

        # フォルダ存在確認・作成
        if not self.fld_pth.exists():
            self.fld_pth.mkdir()

        # 保存して閉じる
        self.wb.save(self.fle_pth)
        self.wb.close()

    def read_workbook(self, rd_fld, rd_fle):
        rd_fld_pth = Path(rd_fld)
        rd_fle_pth = rd_fld_pth.joinpath(rd_fle)        
        self.wb = px.load_workbook(rd_fle_pth)
        self.ws = self.wb['Ranking']
        return self.ws
